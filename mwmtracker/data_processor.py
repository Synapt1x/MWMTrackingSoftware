# -*- coding: utf-8 -*-

""" Code used for processing the tracking data from the mouse tracker,
specifically by assembling a pandas dataframe and writing this to an excel
using XlsxWriter from pandas.

data_processor.py: this file contains the code implementing the functionality
for saving the output path data to an excel file.

"""


import os

import pandas as pd
import numpy as np
import cv2
import util
import openpyxl

# imports for statistics
from scipy.stats import ttest_ind, ttest_rel
import statsmodels.api as sm
from statsmodels.formula.api import ols
from statsmodels.stats.anova import anova_lm
from statsmodels.graphics.factorplots import interaction_plot


class Data_processor:
    """
    Data Processor class
    """

    def __init__(self, excelFilename='trackingData.xlsx',
                 dataFilename='PrelimData.xlsx', cm_scale=6.0):
        """constructor"""

        self.excelFilename = excelFilename
        self.dataFilename = dataFilename
        self.initialize_writer()

        # init data
        self.columns = ['vid num', 'x', 'y', 'dist', 't']
        self.dist_columns = ['vid num', 'dist']
        self.time_dist_columns = ['vid num', 'dist', 'Time']
        self.output_data = pd.DataFrame(columns=self.columns)
        self.dist_data = pd.DataFrame(columns=self.dist_columns)
        self.tracking_data = pd.DataFrame(columns=self.time_dist_columns)

        if not os.path.exists(self.excelFilename):
            dirname = os.path.dirname(self.excelFilename)
            if not os.path.exists(dirname):
                os.makedirs(dirname)

            # write to excels to initialize them
            self.output_data.to_excel(self.excelWriter, 'All Tracking Data')
            self.dist_data.to_excel(self.excelWriter, 'Dist Data')
            self.tracking_data.to_excel(self.excelWriter,
                                        sheet_name="Timed Dist Data")

            self.excelWriter.save()

        self.dist_data = pd.read_excel(self.excelFilename,
                                       sheet_name="Timed Dist Data")

        self.cm_scale = cm_scale

    def initialize_writer(self):
        self.excelWriter = pd.ExcelWriter(self.excelFilename,
                                          engine='xlsxwriter')

    def save_frame(self, col, x_locs, y_locs, t):
        """
        Update Pandas dataframe with current frame information.

        :param col:
        :param x_locs:
        :param y_locs:
        :param t:
        :return:
        """

        col = int(col.split(" ")[-1])

        vid_name = np.repeat(col, len(x_locs))

        # store all data from x, y locs through tracking
        temp_df = pd.DataFrame(columns=['vid num', 'x', 'y', 'dist', 't'])

        temp_df['vid num'] = vid_name
        temp_df['x'] = x_locs
        temp_df['y'] = y_locs
        x_diff = np.power(temp_df['x'] - temp_df['x'].shift(1), 2)
        y_diff = np.power(temp_df['y'] - temp_df['y'].shift(1), 2)
        dists = np.sqrt(x_diff + y_diff) / self.cm_scale
        temp_df['dist'] = np.nancumsum(dists)
        temp_df['t'] = t

        # store all total dist data for each tracking effort
        dist_df = pd.DataFrame(columns=['vid num', 'dist'])
        dist_df = dist_df.append(pd.Series([col, temp_df.iloc[-1]['dist']],
                                 index=dist_df.columns), ignore_index=True)

        self.output_data = pd.concat([self.output_data, temp_df])
        self.dist_data = pd.concat([self.dist_data, dist_df])

        self.save_to_excel()

    def save_to_excel(self):
        """
        Save to the
        :return:
        """

        self.initialize_writer()

        self.output_data.to_excel(self.excelWriter, 'All Tracking Data')
        self.dist_data.to_excel(self.excelWriter, 'Dist Data')

        self.excelWriter.save()
        self.excelWriter.close()

    def no_collapse_df(self, df, groups, output_var):

        out_df = df.groupby(['Day', 'Group'])['Dist'].mean()
        dfStd = df.groupby(['Day', 'Group'])['Dist'].std()
        dfSem = df.groupby(['Day', 'Group'])[
            'Dist'].sem()
        dfStd.name = 'std'
        dfSem.name = 'sem'
        final_df = pd.concat([df, dfStd, dfSem], axis=1).reset_index()

    def collapse_df(self, df, groups, output_var, columns):

        final_df = pd.DataFrame(columns=columns)

        out_df = df.groupby(groups)[output_var].mean().reset_index()
        out_df_std = df.groupby(groups)[output_var].std().reset_index()
        out_df_sem = df.groupby(groups)[output_var].sem().reset_index()

        control_out_df = out_df.where(out_df['group'] ==
                                      'Control').dropna().reset_index()
        exp_out_df = out_df.where(out_df['group'] ==
                                  'Nilotinib').dropna().reset_index()

        control_out_std = out_df_std.where(out_df['group'] ==
                                           'Control').dropna().reset_index()
        exp_out_std = out_df_std.where(out_df['group'] ==
                                       'Nilotinib').dropna().reset_index()

        control_out_sem = out_df_sem.where(out_df['group'] ==
                                           'Control').dropna().reset_index()
        exp_out_sem = out_df_sem.where(out_df['group'] ==
                                       'Nilotinib').dropna().reset_index()

        final_df[columns[0]] = control_out_df['day']
        final_df[columns[1]] = control_out_df[output_var]
        final_df[columns[2]] = exp_out_df[output_var]
        final_df[columns[3]] = control_out_std[output_var]
        final_df[columns[4]] = exp_out_std[output_var]
        final_df[columns[5]] = control_out_sem[output_var]
        final_df[columns[6]] = exp_out_sem[output_var]

        return final_df

    def merge_tracking_data_times(self, vid_folder):

        time_df = self.write_raw_times(vid_folder)

        new_writer = pd.ExcelWriter(self.excelFilename, engine='openpyxl')

        if os.path.exists(self.excelFilename):
            book = openpyxl.load_workbook(self.excelFilename)
            new_writer.book = book

        self.tracking_data = pd.read_excel(self.excelFilename, 'Dist Data')

        self.tracking_data = self.tracking_data.merge(time_df,
                                                      how='inner',
                                                      on='vid num')
        self.tracking_data.sort_values(['vid num'], inplace=True)
        self.tracking_data = self.tracking_data.reset_index(drop=True)

        self.tracking_data.to_excel(new_writer, 'Timed Dist Data')
        new_writer.save()
        new_writer.close()

    def fix_prelim_data(self):

        self.output_data = pd.read_excel(self.dataFilename)
        self.tracking_data = pd.read_excel(self.excelFilename,
                                           'Timed Dist Data')

        self.add_swim_speed()
        self.add_mouse_ids()

        new_writer = pd.ExcelWriter(self.excelFilename,
                                    engine='openpyxl')
        if os.path.exists(self.excelFilename):
            book = openpyxl.load_workbook(self.excelFilename)
            new_writer.book = book

        self.tracking_data.to_excel(new_writer,
                                    sheet_name='Complete Distance Data')
        new_writer.save()
        new_writer.close()

    def add_swim_speed(self):

        self.tracking_data['swim speed'] = self.tracking_data['dist'] / \
                                           self.tracking_data['Time']

    def create_group_comparisons(self, df, writer, close=True,
                                 overwrite=False, df_name=None):

        day_latency = self.collapse_df(df, ['day', 'group'],
                                       'Time', ['Day', 'Control Time',
                                                'Nilotinib Time',
                                                'Control std',
                                                'Nilotinib std',
                                                'Control sem',
                                                'Nilotinib sem'])
        day_dists = self.collapse_df(df, ['day', 'group'],
                                     'dist', ['Day', 'Control Dist',
                                              'Nilotinib Dist',
                                              'Control std',
                                              'Nilotinib std',
                                              'Control sem',
                                              'Nilotinib sem'])
        day_speeds = self.collapse_df(df, ['day', 'group'],
                                      'swim speed', ['Day', 'Control '
                                                            'Swim Speed',
                                                     'Nilotinib Swim Speed',
                                                     'Control std',
                                                     'Nilotinib std',
                                                     'Control sem',
                                                     'Nilotinib sem'])

        if not overwrite:
            if os.path.exists(df_name):
                book = openpyxl.load_workbook(df_name)
                writer.book = book

        day_latency.to_excel(writer, 'Latency by Day')
        day_dists.to_excel(writer, 'Path Length by Day')
        day_speeds.to_excel(writer, 'Swim Speed by Day')

        writer.save()
        if close:
            writer.close()

    def run_anovas(self):

        data_filename = os.sep.join(self.excelFilename.split(os.sep)[:-1]) \
                          + os.sep + 'AllFinalTrackingData.xlsx'

        data = pd.read_excel(data_filename,
                             sheet_name='Complete Distance Data')
        memory_data = pd.read_excel(data_filename,
                             sheet_name='Memory Data')
        tracking_writer = pd.ExcelWriter(data_filename,
                                         engine="openpyxl")

        old_data = data.copy()
        old_memory_data = memory_data.copy()

        # add additional sheets with dist, latency, and swim speed comparisons
        self.create_group_comparisons(data, tracking_writer, close=False,
                                        df_name=data_filename)

        ##========= mouse number to be excluded from analysis ============ ##
        exclude_on = True
        while exclude_on:
            exclude_on = input("Do you wish to exclude any particular mouse from analysis? Y/N\n")
            if exclude_on.lower() in {'y', 'yes'}:
                exclude_num = input('Which mouse would you like to exclude?\n')
                try:
                    data = old_data.copy()
                    memory_data = old_memory_data.copy()

                    exclude_num = int(exclude_num)
                    assert(exclude_num < 21 and exclude_num > 0)

                    data = data.loc[data['mouse'] != exclude_num]
                    memory_data = memory_data.loc[memory_data['mouse'] != exclude_num]

                    print(f"****** RESULTS WITHOUT MOUSE {exclude_num} ******")

                    formula = 'dist ~ group * day * trial + (1|mouse)'
                    model = ols(formula, data).fit()
                    dist_aov_table = anova_lm(model, typ=2)
                    print("Learning three-way ANOVA for path length:")
                    print(dist_aov_table)

                    formula = 'Time ~ group * day * trial + (1|mouse)'
                    model = ols(formula, data).fit()
                    time_aov_table = anova_lm(model, typ=2)
                    print("Learning three-way ANOVA for escape latency:")
                    print(time_aov_table)

                    mem_formula = 'ACI ~ Group * trial + (1|mouse)'
                    mem_model = ols(mem_formula, memory_data).fit()
                    mem_aov_table = anova_lm(mem_model, typ=2)
                    print("Memory two-way ANOVA for annulus crossing index:")
                    print(mem_aov_table)

                    new_df = memory_data.rename(columns={'time target proportion':
                                                            'targetProp'})
                    mem_formula = 'targetProp ~ Group * trial + (1|mouse)'
                    mem_model = ols(mem_formula, new_df).fit()
                    mem_time_target_aov_table = anova_lm(mem_model, typ=2)
                    print("Memory two-way ANOVA for proportion in target quadrant:")
                    print(mem_time_target_aov_table)
                except:
                    print("Not a valid number!")
            else:
                data = old_data.copy()
                memory_data = old_memory_data.copy()

                print("\n********** FULL RESULTS ***********")

                formula = 'dist ~ group * day * trial + (1|mouse)'
                model = ols(formula, data).fit()
                dist_aov_table = anova_lm(model, typ=2)
                print("Learning three-way ANOVA for path length:")
                print(dist_aov_table)

                formula = 'Time ~ group * day * trial + (1|mouse)'
                model = ols(formula, data).fit()
                time_aov_table = anova_lm(model, typ=2)
                print("Learning three-way ANOVA for escape latency:")
                print(time_aov_table)

                mem_formula = 'ACI ~ Group * trial + (1|mouse)'
                mem_model = ols(mem_formula, memory_data).fit()
                mem_aov_table = anova_lm(mem_model, typ=2)
                print("Memory two-way ANOVA for annulus crossing index:")
                print(mem_aov_table)

                new_df = memory_data.rename(columns={'time target proportion':
                                                        'targetProp'})
                mem_formula = 'targetProp ~ Group * trial + (1|mouse)'
                mem_model = ols(mem_formula, new_df).fit()
                mem_time_target_aov_table = anova_lm(mem_model, typ=2)
                print("Memory two-way ANOVA for proportion in target quadrant:")
                print(mem_time_target_aov_table)

                break

    def add_mouse_ids(self):

        skip_videos = [6, 95, 98, 143, 263]
        has_double = [83, 104, 123, 144, 147, 161, 180, 239, 319, 398]
        day_orders = {1: [1, 11, 2, 12, 3, 13, 4, 14, 5, 15,
                          6, 16, 7, 17, 8, 18, 9, 19, 10, 20],
                      2: [2, 12, 3, 13, 8, 18, 4, 14, 7, 17,
                          10, 20, 1, 11, 9, 19, 6, 16, 5, 15],
                      3: [9, 19, 8, 18, 1, 11, 4, 14, 6, 16,
                          3, 13, 5, 15, 10, 20, 2, 12, 7, 17],
                      4: [1, 11, 10, 20, 2, 12, 9, 19, 3, 13,
                          4, 14, 8, 18, 7, 17, 6, 16, 5, 15],
                      5: [4, 14, 6, 16, 3, 13, 7, 17, 5, 15,
                          9, 19, 2, 12, 1, 11, 10, 20, 8, 18],
                      6: [1, 11, 2, 12, 3, 13, 4, 14, 5, 15,
                          6, 16, 7, 17, 8, 18, 9, 19, 10, 20]}
        groups = {1: 'Nilotinib', 2: 'Nilotinib', 3: 'Nilotinib', 4: 'Nilotinib',
                  5: 'Nilotinib', 6: 'Nilotinib', 7: 'Nilotinib', 8: 'Nilotinib',
                  9: 'Nilotinib', 10: 'Nilotinib',
                  11: 'Control', 12: 'Control', 13: 'Control',
                  14: 'Control', 15: 'Control', 16: 'Control',
                  17: 'Control', 18: 'Control', 19: 'Control',
                  20: 'Control'}
        video_order = [x for x in range(3, 478) if x not in skip_videos]
        for add_double in has_double:
            i_vid = video_order.index(add_double)
            video_order.insert(i_vid + 1, add_double * 10)

        num_days = 6
        num_trials = 4
        num_vids = len(video_order)

        vid_num = 0
        done = False

        self.tracking_data['mouse'] = np.nan
        self.tracking_data['group'] = ''
        self.tracking_data['day'] = 0
        self.tracking_data['trial'] = 0

        for day in range(1, num_days + 1):

            if done:
                break

            trial = 1
            while trial < num_trials + 1:

                if done:
                    break

                for mouse in day_orders[day]:

                    video = video_order[vid_num]

                    group = groups[mouse]

                    vid_num += 1

                    vid_idx = self.tracking_data['vid num'] == video

                    self.tracking_data.loc[vid_idx, 'mouse'] = mouse
                    self.tracking_data.loc[vid_idx, 'group'] = group
                    self.tracking_data.loc[vid_idx, 'day'] = day
                    self.tracking_data.loc[vid_idx, 'trial'] = trial

                trial += 1

        self.tracking_data.dropna(inplace=True)

    def compute_annulus_crossing_index(self, target_bounds, quadrants):

        target_min_x, target_min_y, target_max_x, target_max_y = target_bounds
        mid_x, mid_y = quadrants

        quad_a_targ = [mid_x + (mid_x - target_max_x),
                       target_min_y, mid_x + (mid_x - target_min_x),
                       target_max_y]
        quad_b_targ = [target_min_x,
                       mid_y + (mid_y - target_max_y),
                       target_max_x, mid_y + (mid_y - target_min_y)]
        quad_c_targ = [mid_x + (mid_x - target_max_x),
                       mid_y + (mid_y - target_max_y),
                       mid_x + (mid_x - target_min_x),
                       mid_y + (mid_y - target_min_y)]

        cols = ['vid num', 'mouse', 'Group', 'trial', 'ACI', 'time target',
                'time target proportion']

        def check_in_bound(x, y):

            return target_min_x < target_max_x and target_min_y < y < \
                target_max_y

        def check_in_other_bound(x, y):

            in_a = quad_a_targ[0] < x < quad_a_targ[2] and quad_a_targ[1] < \
                   y < quad_a_targ[3]
            in_b = quad_b_targ[0] < x < quad_b_targ[2] and quad_b_targ[1] < \
                   y < quad_b_targ[3]
            in_c = quad_c_targ[0] < x < quad_c_targ[2] and quad_c_targ[1] < \
                   y < quad_c_targ[3]

            return in_a or in_b or in_c

        output_df = pd.DataFrame(columns=cols)

        probe_data = pd.read_excel(self.excelFilename,
                                          'Probe Tracking Data')

        vid_nums, num_counts = np.unique(probe_data['vid num'].values,
                                         return_counts=True)

        mouse_nums = [1, 11, 2, 12, 3, 13, 4, 14, 5, 15, 6, 16, 7, 17,
                      8, 18, 9, 19, 10, 20]
        groups = ['Nilotinib', 'Control', 'Nilotinib', 'Control', 'Nilotinib',
                  'Control', 'Nilotinib', 'Control', 'Nilotinib', 'Control',
                  'Nilotinib', 'Control', 'Nilotinib', 'Control',
                  'Nilotinib', 'Control', 'Nilotinib', 'Control',
                  'Nilotinib', 'Control']
        trials = [1, 2, 3, 4]

        m_num = 0
        t_num = 0

        for vid in vid_nums:

            mouse = mouse_nums[m_num]
            group = groups[m_num]
            trial = trials[t_num]

            if m_num == len(mouse_nums) - 1:
                t_num += 1
            m_num = (m_num + 1) % len(mouse_nums)

            passes = 0
            passes_other = 0
            time_target = 0.
            time_other = 0.

            prev_x = 0
            prev_y = 0
            prev_t = 0.

            temp_df = probe_data[probe_data['vid num'] == vid]

            for i, row in temp_df.iterrows():
                x = row['x']
                y = row['y']
                dt = row['Time'] - prev_t

                if check_in_bound(x, y) and not check_in_bound(prev_x, \
                        prev_y):
                    passes += 1
                if check_in_other_bound(x, y) and not check_in_other_bound(
                    prev_x, prev_y):
                    passes_other += 1

                if x < mid_x and y < mid_y:
                    time_target += dt
                else:
                    time_other += dt

                prev_t = row['Time']
                prev_x = x
                prev_y = y

            if passes_other > 1E-12:
                aci = float(passes) / passes_other
            else:
                aci = 0.
            if time_other > 1E-12:
                time_target_proportion = float(time_target) / time_other
            else:
                time_target_proportion = 0.

            output_df = output_df.append(pd.Series(
                [vid, mouse, group, trial, aci, time_target,
                 time_target_proportion], index=output_df.columns),
                ignore_index=True)

        new_writer = pd.ExcelWriter(self.excelFilename, engine='openpyxl')

        memory_filename = os.sep.join(self.excelFilename.split(os.sep)[:-1]) \
                          + os.sep + 'MemoryData.xlsx'
        single_writer = pd.ExcelWriter(memory_filename, engine='xlsxwriter')

        if os.path.exists(self.excelFilename):
            book = openpyxl.load_workbook(self.excelFilename)
            new_writer.book = book

        group_means = output_df.groupby('Group').mean()
        group_std = output_df.groupby('Group').std()
        group_stderr = output_df.groupby('Group')['ACI', 'time target',
                                                  'time target proportion',
                                                  ].sem()
        group_std = group_std.rename(
            columns={'ACI': 'ACI std', 'time target': 'time target std',
                     'time target proportion': 'time target proportion std'})
        group_stderr = group_stderr.rename(
            columns={'ACI': 'ACI sem', 'time target': 'time target sem',
                     'time target proportion': 'time target proportion sem'})

        group_df = group_means.merge(group_std, how='outer', left_index=True,
                                     right_index=True)
        group_df = group_df.merge(group_stderr, how='outer', left_index=True,
                                     right_index=True)

        output_df.to_excel(new_writer, sheet_name='Probe Data')
        group_df.to_excel(new_writer, sheet_name='Probe Comparison')

        aci_t_val, aci_p_val = ttest_ind(
            *output_df.groupby('Group')['ACI'].apply(lambda x: list(x)),
            equal_var=False)

        time_t_val, time_p_val = ttest_ind(
            *output_df.groupby('Group')['time target'].apply(lambda x: list(x)),
            equal_var=False)

        time_prop_t_val, time_prop_p_val = ttest_ind(
            *output_df.groupby('Group')['time target proportion'].apply(lambda
                                                                            x: list(x)),
            equal_var=False)

        # write stats to excel and save main excel
        stats_df = pd.DataFrame()
        stats_df['ACI p-val'] = [aci_p_val]
        stats_df['time target p-val'] = [time_p_val]
        stats_df['time target proportion p-val'] = [time_prop_p_val]

        stats_df.to_excel(new_writer, sheet_name='Probe Stats')
        new_writer.save()
        new_writer.close()

        # write all data to single memory sheet as well
        group_df.to_excel(single_writer, sheet_name='Probe Comparison')
        stats_df.to_excel(single_writer, sheet_name='Probe Stats')
        output_df.to_excel(single_writer, sheet_name='Probe Data')
        probe_data.to_excel(single_writer, sheet_name='All Probe Tracking Data')

        single_writer.save()
        single_writer.close()

    def write_raw_times(self, vid_folder):

        raw_times_df = pd.DataFrame(columns=['vid num', 'Time'])

        all_files = util.load_files(vid_folder)

        for video_name in all_files:

            vid_num = int(video_name.split(" ")[-1].split(".")[0])

            vid = cv2.VideoCapture(video_name)
            valid, _ = vid.read()

            if valid:
                if int(vid.get(cv2.CAP_PROP_FPS)) == 0:
                    time = 0
                    found = False
                else:
                    time = int(vid.get(cv2.CAP_PROP_FRAME_COUNT)) / int(
                        vid.get(cv2.CAP_PROP_FPS))
                    if time >= 90:
                        time = 90
                        found = False
                    else:
                        found = True

            raw_times_df = raw_times_df.append(pd.Series([vid_num, time],
                                               index=raw_times_df.columns),
                                               ignore_index=True)

        raw_times_df['vid num'] = raw_times_df['vid num'].astype(int)

        return raw_times_df

    def write_data(self, data_folder=None, num_days=6, num_trials=4, n=10,
                  num_vids=480):
        """
        Save video IDs to a file

        :return:
        """

        by_individual = False
        without_five = False

        self.output_data = pd.read_excel(self.dataFilename)
        self.dist_data = pd.read_excel(self.dataFilename,
                                       sheet_name="Complete Dist Data")

        self.output_data.dropna(inplace=True)
        self.output_data['Found'] = self.output_data['Time'] != 90
        self.output_data['Group'] = self.output_data['ID'].apply(lambda x: 'Nilotinib' if x in
                                                                        range(1, 11) else 'Control')

        # base_df = self.add_tracking_data(base_df)
        self.all_data = self.output_data.copy()

        writer = pd.ExcelWriter(data_folder + os.sep + 'data.xlsx',
                                engine='xlsxwriter')
        self.output_data.sort_values(['Day', 'Trial', 'ID'], inplace=True)
        self.output_data.to_excel(writer, 'All Data')

        trial_data = self.output_data.set_index('Trial')
        trial_data = trial_data.groupby(['Day', 'ID',
                                         'Group']).mean().reset_index()
        trial_data.to_excel(writer, 'Data by trial')

        if without_five:

            other_df = self.output_data.groupby(['Day', 'Group']).mean()
            other_df = other_df.drop(columns=['ID', 'Trial', 'Found'])

            other_df.to_excel(writer, 'Without Five Trial Averages')

        if by_individual:

            day_latency = self.collapse_df(self.output_data, ['Day', 'Group'],
                                           'Time', ['Day', 'Control Time',
                                                    'Nilotinib Time',
                                                    'Control std',
                                                    'Nilotinib std',
                                                    'Control sem',
                                                    'Nilotinib sem'])
            day_dists = self.collapse_df(self.output_data, ['Day', 'Group'],
                                         'Dist', ['Day', 'Control Dist',
                                                  'Nilotinib Dist',
                                                  'Control std',
                                                  'Nilotinib std',
                                                  'Control sem',
                                                  'Nilotinib sem'])
            day_speeds = self.collapse_df(self.output_data, ['Day', 'Group'],
                                         'Swim Speed', ['Day', 'Control '
                                                               'Swim Speed',
                                                        'Nilotinib Swim Speed',
                                                        'Control std',
                                                        'Nilotinib std',
                                                        'Control sem',
                                                        'Nilotinib sem'])
        else:
            day_latency = self.collapse_df(trial_data, ['Day', 'Group'],
                                           'Time', ['Day', 'Control Time',
                                                    'Nilotinib Time',
                                                    'Control std',
                                                    'Nilotinib std',
                                                    'Control sem',
                                                    'Nilotinib sem'])
            day_dists = self.collapse_df(trial_data, ['Day', 'Group'],
                                           'Dist', ['Day', 'Control Dist',
                                                    'Nilotinib Dist',
                                                    'Control std',
                                                    'Nilotinib std',
                                                    'Control sem',
                                                    'Nilotinib sem'])
            day_speeds = self.collapse_df(trial_data, ['Day', 'Group'],
                                           'Swim Speed', ['Day', 'Control '
                                                                 'Swim Speed',
                                                    'Nilotinib Swim Speed',
                                                    'Control std',
                                                    'Nilotinib std',
                                                    'Control sem',
                                                    'Nilotinib sem'])

        day_latency.to_excel(writer, 'Latency by Day')
        day_dists.to_excel(writer, 'Path Length by Day')
        day_speeds.to_excel(writer, 'Swim Speed by Day')

        writer.save()
        writer.close()

        return self.output_data


if __name__ == '__main__':
    print("Please run the file 'main.py'")
